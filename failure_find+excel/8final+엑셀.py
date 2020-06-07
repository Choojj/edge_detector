import cv2
import numpy as np
import openpyxl

def extraction(path): # 원본에서 roi 추출
    image = cv2.imread(path, cv2.IMREAD_COLOR)

    # blurred = cv2.GaussianBlur(image, (5, 5), 0)

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) 

    edge = cv2.Canny(gray, 50, 150)

    edge = cv2.bitwise_not(edge)

    contours = cv2.findContours(edge.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    cv2.drawContours(edge, contours[0], -1, (0, 0, 0), 1)

    nlabels, labels, stats, centroids = cv2.connectedComponentsWithStats(edge)
    for i in range(nlabels):
        if i < 2:
            continue

        area = stats[i, cv2.CC_STAT_AREA]
        center_x = int(centroids[i, 0])
        center_y = int(centroids[i, 1]) 
        left = stats[i, cv2.CC_STAT_LEFT]
        top = stats[i, cv2.CC_STAT_TOP]
        width = stats[i, cv2.CC_STAT_WIDTH]
        height = stats[i, cv2.CC_STAT_HEIGHT]

        if area > 80000 and area < 1000000:
            dst = image.copy()
            dst = image[top - 10:top + height + 10, left - 10:left + width + 10]

    return dst

def diff_clip(imageA, imageB): # roi에서 차영상 추출
    height, width = 600, 1000
    blank_image = np.zeros((height, width, 3), np.uint8)
    blank_imageA = cv2.bitwise_not(blank_image)
    blank_imageB = cv2.bitwise_not(blank_image)

    grayA = cv2.cvtColor(imageA, cv2.COLOR_BGR2GRAY)
    grayB = cv2.cvtColor(imageB, cv2.COLOR_BGR2GRAY)
    blank_imageA = cv2.cvtColor(blank_imageA, cv2.COLOR_BGR2GRAY)
    blank_imageB = cv2.cvtColor(blank_imageB, cv2.COLOR_BGR2GRAY)

    ret, resultA = cv2.threshold(grayA, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    ret, resultB = cv2.threshold(grayB, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)

    resultA_y, resultA_x = resultA.shape
    resultB_y, resultB_x = resultB.shape

    blank_imageA[height - resultA_y:height, 0:resultA_x] = resultA
    blank_imageB[height - resultB_y:height, 0:resultB_x] = resultB

    absdiff = cv2.absdiff(blank_imageA, blank_imageB)
    return absdiff, blank_imageA, blank_imageB

#ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ

tolerance = 2
original_x = [138, 268, 398, 528]
original_y = [202, 460]
original_height, original_width = 258, 130
failure_mode = 0 # 0양품 1길이높이감소 2길이높이증가 3길이폭감소 4길이폭증가 5균일함몰 6균일돌출 7비균일함몰 8비균일돌출 9채움 10구멍

count1 = 0 # 차도형꼭지점과 원도형꼭지점의 일치개수
count2 = 0 # 대칭도형의 개수
count3 = 0 # 현재 도형이 구멍인가
count4 = 0 # 모든 도형이 구멍인가
count5 = 0 # 채움 확인
count6 = 0 # 길이 높이 확인
count7 = 0 # 길이 폭 확인
count8 = 0 # 현재 도형이 함몰인가
count9 = 0 # 모든 도형이 함몰인가
condition1 = condition2 = condition3 = bool # 균일판별 3조건
uniform = bool # 균일 비균일
inner = bool # 함몰 돌출

vertex_location = []
min_max_location = []
center_location = []

path1 = "C:/Users/HSWB/Desktop/edge_detector/data1/original_01.png"
path2 = "C:/Users/HSWB/Desktop/edge_detector/data1/modified_02.png" # 파일위치

clip1 = extraction(path1)
clip2 = extraction(path2) # 파일 roi 추출

absdiff, blankA, blankB = diff_clip(clip1, clip2) # roi 차영상 생성

contours, hierarchy = cv2.findContours(absdiff, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

for cnt in enumerate(contours): # 차영상에서 도형 정보 수집
    epsilon = 0.005 * cv2.arcLength(cnt[1], True)
    approx = cv2.approxPolyDP(cnt[1], epsilon, True)
    size = len(approx)

    vertex_location.append([]) # 도형 꼭지점 좌표
    for i in range(size):
        vertex_location[cnt[0]].append(approx[i][0])

    min_location_x = float("inf")
    max_location_x = float("-inf")
    min_location_y = float("inf")
    max_location_y = float("-inf")

    for i in range(size): # 경계사각형 꼭지점 좌표
        if approx[i][0][0] > max_location_x:
            max_location_x = approx[i][0][0]
        if approx[i][0][0] < min_location_x:
            min_location_x = approx[i][0][0]
        if approx[i][0][1] > max_location_y:
            max_location_y = approx[i][0][1]
        if approx[i][0][1] < min_location_y:
            min_location_y = approx[i][0][1]
    
    min_max_location.append([])
    min_max_location[cnt[0]].append(min_location_x)
    min_max_location[cnt[0]].append(min_location_y)
    min_max_location[cnt[0]].append(max_location_x)
    min_max_location[cnt[0]].append(max_location_y)
    min_max_location[cnt[0]].append(size)

    mmt = cv2.moments(cnt[1]) # 도형의 무게중심 좌표
    center_x = int(mmt["m10"] / mmt["m00"])
    center_y = int(mmt["m01"] / mmt["m00"])

    center_location.append([])
    center_location[cnt[0]].append(center_x)
    center_location[cnt[0]].append(center_y)

    print(cnt[0], size, center_x, center_y, end=" ")
    for i in range(size):
        print(approx[i][0], end=" ")
    print()

print(f"min_max_location = {min_max_location}")
print(f"center_location = {center_location}")
print()

if (len(contours) == 0):
    print("양품")
    failure_mode = 0
else:
    if len(contours) % 2 == 0: # 도형이 쌍으로 존재하는가?
        print(f"{len(contours)}, condition1 yes")
        condition1 = True
    else:
        print(f"{len(contours)}, condition1 no")
        condition1 = False
    print()

    for i in range(len(contours)): # 도형의 꼭지점의 원본 꼭지점과 2개이상 일치하는가?
        if (condition1 == False):
            print("condition2 no")
            condition2 = False
            break

        for j in range(len(vertex_location[i])):
            print(j, vertex_location[i][j][0], vertex_location[i][j][1])
            if ((original_x[0] - tolerance <= vertex_location[i][j][0] <= original_x[0] + tolerance
                or original_x[1] - tolerance <= vertex_location[i][j][0] <= original_x[1] + tolerance
                or original_x[2] - tolerance <= vertex_location[i][j][0] <= original_x[2] + tolerance
                or original_x[3] - tolerance <= vertex_location[i][j][0] <= original_x[3] + tolerance)
                and (original_y[0] - tolerance - original_height * 0.1 <= vertex_location[i][j][1] <= original_y[0] + tolerance + original_height * 0.1
                or original_y[1] - tolerance - original_height * 0.1 <= vertex_location[i][j][1] <= original_y[1] + tolerance + original_height * 0.1)):
                print(j, vertex_location[i][j][0], vertex_location[i][j][1])
                count1 += 1
        if (count1 >= 2):
            print(f"{count1}, condition2 yes")
            condition2 = True
            count1 = 0
        else:
            print(f"{count1}, condition2 no")
            condition2 = False
            count1 = 0
            break
    print()

    for i in range(0, len(min_max_location), 2): # 도형쌍 끼리 대칭인가?
        if (condition2 == False):
            condition3 = False
            break
        print(min_max_location[i], min_max_location[i + 1])
        if (min_max_location[i][0] - min_max_location[i + 1][0] - tolerance <= min_max_location[i][2] - min_max_location[i + 1][2] <= min_max_location[i][0] - min_max_location[i + 1][0] + tolerance
            and min_max_location[i][1] - tolerance <= min_max_location[i + 1][1] <= min_max_location[i][1] + tolerance
            and min_max_location[i][3] - tolerance <= min_max_location[i + 1][3] <= min_max_location[i][3] + tolerance):
            count2 += 1

    if (count2 == len(min_max_location) / 2):
        print("condittion3 yes")
        condition3 = True
    else:
        print("condition3 no")
        condition3 = False
    count2 = 0

    for i in range(len(contours)): # 도형 무게중심이 원본 내부에 있는가?
        if ((original_x[0] + tolerance < center_location[i][0] < original_x[1] - tolerance
            or original_x[2] + tolerance < center_location[i][0] < original_x[3] - tolerance)
            and original_y[0] + tolerance < center_location[i][1] < original_y[1] - tolerance):
            print(center_location[i][0], center_location[i][1])
            count8 += 1
    if (count8 == len(contours)):
        print("함몰")
        inner = True
    else:
        print("돌출")
        inner = False
    count8 = 0

    print()

    if (condition1 == condition2 == condition3 == True): # 균일 판별 조건을 모두 만족하면 균일
        print("균일")
        uniform = True
    else:
        print("비균일")
        uniform = False

    if (uniform == True):
        if (inner == True):
            print("균일 함몰")
            failure_mode = 5
        else:
            print("균일 돌출")
            failure_mode = 6

        for i in range(len(contours)): # 폭이 원본과 같으면
            if (original_width - 2 <= abs(min_max_location[i][0] - min_max_location[i][2]) <= original_width + 2 and min_max_location[i][4] == 4):
                print(abs(min_max_location[i][0] - min_max_location[i][2]), original_width)
                count6 += 1
        if (count6 == len(contours)):
            print("길이 높이")
            count6 = 0
            if (inner == True):
                print("길이 높이 감소")
                failure_mode = 1
            else:
                print("길이 높이 증가")
                failure_mode = 2
        
        for i in range(len(contours)): # 높이가 원본과 같으면
            if (original_height - 2 <= abs(min_max_location[i][1] - min_max_location[i][3]) <= original_height + 2 and min_max_location[i][4] == 4):
                print(abs(min_max_location[0][1] - min_max_location[0][3]), original_height)
                count6 += 1
        if (count6 == len(contours)):
            print("길이 폭")
            count6 = 0
            if (inner == True):
                print("길이 폭 감소")
                failure_mode = 3
            else:
                print("길이 폭 증가")
                failure_mode = 4

    else:
        if (inner == True):
            print("비균일 함몰")
            failure_mode = 7
        else:
            print("비균일 돌출")
            failure_mode = 8

        for i in range(len(contours)): # 채움조건
            for j in range(len(vertex_location[i])):
                if ((original_x[1] - tolerance <= vertex_location[i][j][0] <= original_x[1] + tolerance
                    or original_x[2] - tolerance <= vertex_location[i][j][0] <= original_x[2] + tolerance)
                    and (original_y[0] - tolerance <= vertex_location[i][j][1] <= original_y[0] + tolerance
                    or original_y[1] - tolerance <= vertex_location[i][j][1] <= original_y[1] + tolerance)):
                    print(j, vertex_location[i][j][0], vertex_location[i][j][1])
                    count5 += 1
            if (count5 == len(vertex_location[i])):
                print(f"{count5}, {len(vertex_location[i])}, 채움")
                failure_mode = 9

        for i in range(len(contours)): # 모든 도형의 꼭지점이 원본 내부에 있는가?
            for j in range(len(vertex_location[i])):
                if ((original_x[0] + tolerance < vertex_location[i][j][0] < original_x[1] - tolerance
                    or original_x[2] + tolerance < vertex_location[i][j][0] < original_x[3] - tolerance)
                    and original_y[0] + tolerance < vertex_location[i][j][1] < original_y[1] - tolerance):
                    print(j, vertex_location[i][j][0], vertex_location[i][j][1])
                    count3 += 1
            if (count3 == len(vertex_location[i])):
                count4 += 1
                count3 = 0
        
        if (count4 == len(contours)):
            print(f"{count3}, {count4}, {len(contours)} 구멍")
            failure_mode = 10
            count4 = 0

print()
print()
if (failure_mode == 0):
    print(f"{failure_mode}, 양품")
elif (failure_mode == 1):
    print(f"{failure_mode}, 길이 높이 감소, -{(min_max_location[0][3] - min_max_location[0][1]) / original_height * 100}%")
elif (failure_mode == 2):
    print(f"{failure_mode}, 길이 높이 증가, +{(min_max_location[0][3] - min_max_location[0][1]) / original_height * 100}%")
elif (failure_mode == 3):
    print(f"{failure_mode}, 길이 폭 감소, -{(min_max_location[0][2] - min_max_location[0][0]) / original_width * 100}%")
elif (failure_mode == 4):
    print(f"{failure_mode}, 길이 폭 증가, +{(min_max_location[0][2] - min_max_location[0][0]) / original_width * 100}%")
elif (failure_mode == 5):
    print(f"{failure_mode}, 균일 함몰")
elif (failure_mode == 6):
    print(f"{failure_mode}, 균일 돌출")
elif (failure_mode == 7):
    print(f"{failure_mode}, 비균일 함몰")
elif (failure_mode == 8):
    print(f"{failure_mode}, 비균일 돌출")
elif (failure_mode == 9):
    print(f"{failure_mode}, 채움")
elif (failure_mode == 10):
    print(f"{failure_mode}, 구멍")

const = 0.6 # 사진크기 너무 커서 줄임
blankA = cv2.resize(blankA, dsize = (0, 0), fx = const, fy = const, interpolation = cv2.INTER_AREA)
blankB = cv2.resize(blankB, dsize = (0, 0), fx = const, fy = const, interpolation = cv2.INTER_AREA)
absdiff = cv2.resize(absdiff, dsize = (0, 0), fx = const, fy = const, interpolation = cv2.INTER_AREA)

cv2.imshow("original", blankA)
cv2.imshow("modified", blankB)
cv2.imshow("absdiff", absdiff)
cv2.waitKey(0)
cv2.destroyAllWindows()

absdiff = cv2.resize(absdiff, dsize = (0, 0), fx = 0.368, fy = 0.368, interpolation = cv2.INTER_AREA)
cv2.imwrite("C:/Users/HSWB/Desktop/edge_detector/data1/absdiff.png", absdiff)

#ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ

workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title = "data"
sheet2 = workbook.create_sheet("result")
sheet1.column_dimensions["A"].width = 27.5
sheet1.column_dimensions["C"].width = 15
sheet1.column_dimensions["D"].width = 15
sheet1.freeze_panes = "A2"
sheet1.row_dimensions[1].border = openpyxl.styles.borders.Border(bottom = openpyxl.styles.borders.Side(border_style = "thick", color="FF000000"))

# curr_sheet.cell(row=current_row, column=1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
# 가운데 정렬

sheet1["A1"] = "absdiff"
sheet1["B1"] = "합 / 불"
sheet1["C1"] = "불량 구분"
sheet1["D1"] = "오차"
sheet1.merge_cells("E1:X1")
sheet1["E1"] = "vertex_location = (x, y)"

# opnepyxl에서는 셀단위로 스타일 입력해야 적용
# sheet1.cell(1, 1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

if (failure_mode == 0):
    sheet1.cell(2, 2).value = "합"
    sheet1.cell(2, 3).value = "정상"
else:
    sheet1.cell(2, 2).value = "불"
    if (failure_mode == 1):
        sheet1.cell(2, 3).value = "길이 높이 감소"
        sheet1.cell(2, 4).value = "-" + str(round((min_max_location[0][3] - min_max_location[0][1]) / original_height * 100, 4)) + "%"
    elif (failure_mode == 2):
        sheet1.cell(2, 3).value = "길이 높이 증가"
        sheet1.cell(2, 4).value = "+" + str(round((min_max_location[0][3] - min_max_location[0][1]) / original_height * 100, 4)) + "%"
    elif (failure_mode == 3):
        sheet1.cell(2, 3).value = "길이 폭 감소"
        sheet1.cell(2, 4).value = "-" + str(round((min_max_location[0][2] - min_max_location[0][0]) / original_width * 100, 4)) + "%"
    elif (failure_mode == 4):
        sheet1.cell(2, 3).value = "길이 폭 증가"
        sheet1.cell(2, 4).value = "+" + str(round((min_max_location[0][2] - min_max_location[0][0]) / original_width * 100, 4)) + "%"
    elif (failure_mode == 5):
        sheet1.cell(2, 3).value = "균일 함몰"
        sheet1.cell(2, 4).value = "X"
    elif (failure_mode == 6):
        sheet1.cell(2, 3).value = "균일 돌출"
        sheet1.cell(2, 4).value = "X"
    elif (failure_mode == 7):
        sheet1.cell(2, 3).value = "비균일 함몰"
        sheet1.cell(2, 4).value = "X"
    elif (failure_mode == 8):
        sheet1.cell(2, 3).value = "비균일 돌출"
        sheet1.cell(2, 4).value = "X"
    elif (failure_mode == 9):
        sheet1.cell(2, 3).value = "채움"
        sheet1.cell(2, 4).value = "X"
    elif (failure_mode == 10):
        sheet1.cell(2, 3).value = "구멍"
        sheet1.cell(2, 4).value = "X"

for i in range(len(vertex_location)):
    for j in range(0, len(vertex_location[i])):
        sheet1.cell(i + 2, j + 5).value = str(vertex_location[i][j][0]) + ", " + str(vertex_location[i][j][1])

image = openpyxl.drawing.image.Image("C:/Users/HSWB/Desktop/edge_detector/data1/absdiff.png")
sheet1.add_image(image, "A2")

workbook.save("C:/Users/HSWB/Desktop/edge_detector/data1/test.xlsx")
