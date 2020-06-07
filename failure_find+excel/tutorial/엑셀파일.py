import openpyxl

workbook = openpyxl.Workbook()

sheet1 = workbook.active
sheet1.title = "data"
sheet2 = workbook.create_sheet("result")

sheet1["A1"] = "도형"
sheet1.merge_cells("B1:K1")
sheet1["B1"] = "좌표"
#sheet1.row_dimensions[1] = Font(Color="FFBB00")
sheet1.row_dimensions["A"] = openpyxl.styles.Font(bold=True)
sheet1.freeze_panes = "A2"

sheet2.cell(row=4, column=1, value=10)
sheet2["A4"].style = "Check Cell"

image = openpyxl.drawing.image.Image("C:/Users/HSWB/Desktop/edge_detector/data/absdiff_01.png")
sheet2.add_image(image, "A10")


workbook.save("C:/Users/HSWB/Desktop/edge_detector/data/test.xlsx")